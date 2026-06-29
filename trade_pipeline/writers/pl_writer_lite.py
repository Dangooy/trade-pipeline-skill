"""
writers/pl_writer_lite.py — Simplified PL generator for public demo

Generates a basic Packing List from OrderModel without requiring the
private pl-gen engine. Uses simplified packing rules:
  - Fixed 25 kg per carton
  - 36 cartons per pallet
  - Net weight from item data, gross weight = net + pallet self-weight

Production environments use the separate pl-gen engine with
customer-specific templates and packing rules.
"""
import math
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .base_writer import BaseWriter, sc as _sc, mc as _mc, brd as _brd

CALIBRI = "Calibri"
DEFAULT_KG_PER_CARTON = 25.0
DEFAULT_CARTONS_PER_PALLET = 36
DEFAULT_PALLET_SELF_WEIGHT_KG = 28.0
DEFAULT_MEASUREMENT_PER_PALLET_M3 = 0.528
HEADER_COLOR = "1F3864"
BORDER_COLOR = "E0E0E0"


def _fnt(size=10, bold=False, color="000000"):
    return Font(name=CALIBRI, size=size, bold=bold, color=color)


def _aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


@dataclass
class PackingLine:
    description: str
    group_key: str
    pcs: float
    net_weight_kg: float
    cartons: int
    kg_per_carton: float


def _compute_packing(
    items,
    kg_per_carton: float = DEFAULT_KG_PER_CARTON,
    cartons_per_pallet: int = DEFAULT_CARTONS_PER_PALLET,
    pallet_self_weight_kg: float = DEFAULT_PALLET_SELF_WEIGHT_KG,
    measurement_per_pallet_m3: float = DEFAULT_MEASUREMENT_PER_PALLET_M3,
) -> tuple[list[PackingLine], dict]:
    lines = []
    total_pcs = 0
    total_net = 0.0
    total_cartons = 0

    for item in items:
        # 总净重优先级：weight_kg → weight_kg_per_piece * qty → kg_mpcs * qty/1000
        # 注意：用 `is not None` 而非 falsy 检查，0.0 是合法重量（赠品/试样）
        nw = item.weight_kg
        if nw is None:
            wpp = getattr(item, "weight_kg_per_piece", None)
            if wpp is not None:
                nw = wpp * item.quantity
            elif item.kg_mpcs is not None:
                nw = item.kg_mpcs * item.quantity / 1000
            else:
                nw = 0

        # 箱数算法优先级：pcs_per_carton（按件） > kg_per_carton_override（覆盖）> 全局 kg_per_carton
        # 注意：用 `is not None` 区分"用户没填"和"用户填了 0"
        pcs_per_ctn = getattr(item, "pcs_per_carton", None)
        kg_override = getattr(item, "kg_per_carton_override", None)

        if pcs_per_ctn is not None and pcs_per_ctn > 0:
            cartons = max(1, math.ceil(item.quantity / pcs_per_ctn))
        else:
            # 0.0 也是合法覆盖（虽然语义可疑，但 None 才表示"用全局默认"）
            eff_kg_per_carton = kg_override if kg_override is not None and kg_override > 0 else kg_per_carton
            cartons = max(1, math.ceil(nw / eff_kg_per_carton)) if nw > 0 else 1

        kg_ctn = round(nw / cartons, 2) if cartons > 0 else 0

        lines.append(PackingLine(
            description=item.description,
            group_key=item.group_key or "",
            pcs=item.quantity,
            net_weight_kg=round(nw, 2),
            cartons=cartons,
            kg_per_carton=kg_ctn,
        ))
        total_pcs += item.quantity
        total_net += nw
        total_cartons += cartons

    total_pallets = max(1, math.ceil(total_cartons / cartons_per_pallet))
    total_gross = round(total_net + total_pallets * pallet_self_weight_kg, 2)
    total_measurement = round(total_pallets * measurement_per_pallet_m3, 2)

    summary = {
        "total_pcs": total_pcs,
        "total_net_weight": round(total_net, 2),
        "total_gross_weight": total_gross,
        "total_cartons": total_cartons,
        "total_pallets": total_pallets,
        "total_measurement_m3": total_measurement,
    }
    return lines, summary


class PackingInfoMissingError(Exception):
    """Raised when items lack weight info needed for PL generation."""

    def __init__(self, missing_items: list):
        self.missing_items = missing_items
        names = ", ".join(f"#{it.no} {it.description[:40]}" for it in missing_items[:3])
        more = f" 等 {len(missing_items)} 项" if len(missing_items) > 3 else ""
        super().__init__(
            f"PL 生成缺少重量信息: {names}{more}。"
            "请补充每个规格的单件重量 (weight_kg) 或每千件重量 (kg_mpcs)，"
            "或在 GUI 中通过装箱向导填写。"
        )


def _check_weight_completeness(items) -> list:
    """Return list of items missing weight data needed for PL.

    v1.1.0: 也认 weight_kg_per_piece（Gateway 收集）和 pcs_per_carton（按件包装）。
    pcs_per_carton 单独存在时仍需 weight 信息，但 weight_kg_per_piece 是新主路径。

    Bug fix (C4): 用 `is not None` 区分 "用户没提供重量数据" 和 "用户提供了 0kg
    （如赠品/试样/包装内附件）"。前者应该报 missing，后者是合法值不应该报。
    """
    missing = []
    for item in items:
        # 有任何一个 weight 字段非 None 都视为已提供（即使值是 0）
        has_weight_source = (
            item.weight_kg is not None
            or getattr(item, "weight_kg_per_piece", None) is not None
            or item.kg_mpcs is not None
        )
        if not has_weight_source:
            missing.append(item)
    return missing


class PLWriterLite(BaseWriter):
    """Simplified PL Writer for public demo.

    v1.1.0 新增 kwargs:
      - allow_missing_weight=True  : 跳过 safety net（demo / 测试用）
      - packing_review=PackingReview : 用 review.pallet 覆盖默认 pallet 配置
    """

    def write(self, output_path: str, **kwargs) -> dict:
        model = self.model
        items = model.items

        # ── Safety net: detect missing weight info before generating empty PL ──
        # Skipped if caller explicitly accepts zero-weight output (e.g. demo with all zeros).
        if not kwargs.get("allow_missing_weight", False):
            missing = _check_weight_completeness(items)
            if missing:
                raise PackingInfoMissingError(missing)

        # ── Packing review override (v1.1.0) ──
        # 如果上层传入 PackingReview（已 apply 到 model），用它的 pallet 配置覆盖
        review = kwargs.get("packing_review")

        packing_cfg = (self.config or {}).get("packing", {})
        # 优先级：review.pallet > config.packing > module defaults
        eff_kg_per_carton = packing_cfg.get("carton_weight_kg", DEFAULT_KG_PER_CARTON)
        eff_cartons_per_pallet = (
            review.pallet.cartons_per_pallet if review and review.pallet
            else packing_cfg.get("cartons_per_pallet", DEFAULT_CARTONS_PER_PALLET)
        )
        eff_pallet_self_weight = (
            review.pallet.self_weight_kg if review and review.pallet
            else packing_cfg.get("pallet_self_weight_kg", DEFAULT_PALLET_SELF_WEIGHT_KG)
        )
        packing_lines, summary = _compute_packing(
            items,
            kg_per_carton=eff_kg_per_carton,
            cartons_per_pallet=eff_cartons_per_pallet,
            pallet_self_weight_kg=eff_pallet_self_weight,
            measurement_per_pallet_m3=packing_cfg.get(
                "measurement_per_pallet_m3", DEFAULT_MEASUREMENT_PER_PALLET_M3
            ),
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "PACKING LIST"

        col_widths = {"A": 6, "B": 40, "C": 10, "D": 12, "E": 12, "F": 14, "G": 14}
        for letter, w in col_widths.items():
            ws.column_dimensions[letter].width = w

        ws.page_setup.paperSize = 9
        ws.page_setup.orientation = "landscape"

        seller_name_cn = self.seller.get("name_cn", "")
        seller_name_en = self.seller.get("name_en", "")
        buyer_name = self.buyer.get("name_en", "") or self.buyer.get("name_ru", "")
        port_loading = model.derived.port_of_loading or "QINGDAO, CHINA"

        R = 1

        # Header
        if seller_name_cn:
            _mc(ws, R, 1, R, 7, value=seller_name_cn,
                font=Font(name="宋体", size=18, bold=True), align=_aln("center"))
            R += 1

        _mc(ws, R, 1, R, 7, value=seller_name_en,
            font=_fnt(14, bold=True), align=_aln("center"))
        R += 1

        R += 1
        _mc(ws, R, 1, R, 7, value="PACKING LIST",
            font=_fnt(16, bold=True), align=_aln("center"))
        R += 1

        R += 1
        _sc(ws, R, 1, value="TO:", font=_fnt(10))
        _sc(ws, R, 2, value=buyer_name, font=_fnt(10, bold=True))
        _sc(ws, R, 6, value="PI No.:", font=_fnt(10))
        _sc(ws, R, 7, value=model.order.pi_number, font=_fnt(10, bold=True))
        R += 1

        _sc(ws, R, 6, value="Date:", font=_fnt(10))
        _sc(ws, R, 7, value=model.order.date, font=_fnt(10))
        R += 1

        _sc(ws, R, 1, value=f"FROM: {port_loading}", font=_fnt(10))
        R += 1

        R += 1

        # Column headers
        headers = ["No.", "Description", "Cartons", "KGS/CTN", "Qty (pcs)", "Net Weight", "Gross Weight"]
        header_fill = PatternFill("solid", fgColor="1F3864")
        for i, h in enumerate(headers, 1):
            c = _sc(ws, R, i, value=h,
                    font=Font(name=CALIBRI, size=10, bold=True, color="FFFFFF"),
                    align=_aln("center", wrap=True))
            c.fill = header_fill
        R += 1

        # Data rows
        prev_group = None
        thin_bottom = _brd(bottom="thin")

        for idx, line in enumerate(packing_lines, 1):
            if line.group_key and line.group_key != prev_group:
                _sc(ws, R, 2, value=line.group_key, font=_fnt(10, bold=True))
                R += 1
                prev_group = line.group_key

            _sc(ws, R, 1, value=idx, font=_fnt(10), align=_aln("center"), border=thin_bottom)
            _sc(ws, R, 2, value=line.description, font=_fnt(10), border=thin_bottom)
            _sc(ws, R, 3, value=line.cartons, font=_fnt(10), align=_aln("right"),
                border=thin_bottom, num_fmt="#,##0")
            _sc(ws, R, 4, value=line.kg_per_carton, font=_fnt(10), align=_aln("right"),
                border=thin_bottom, num_fmt="#,##0.00")
            _sc(ws, R, 5, value=line.pcs, font=_fnt(10), align=_aln("right"),
                border=thin_bottom, num_fmt="#,##0")
            _sc(ws, R, 6, value=line.net_weight_kg, font=_fnt(10), align=_aln("right"),
                border=thin_bottom, num_fmt="#,##0.00")
            _sc(ws, R, 7, value="", font=_fnt(10), align=_aln("right"), border=thin_bottom)
            R += 1

        # Total row
        total_border = _brd(top="thin", bottom="thin")
        _sc(ws, R, 1, value="", border=total_border)
        _sc(ws, R, 2, value="TOTAL:", font=_fnt(11, bold=True), border=total_border)
        _sc(ws, R, 3, value=summary["total_cartons"], font=_fnt(11, bold=True),
            align=_aln("right"), border=total_border, num_fmt="#,##0")
        _sc(ws, R, 4, value="", border=total_border)
        _sc(ws, R, 5, value=summary["total_pcs"], font=_fnt(11, bold=True),
            align=_aln("right"), border=total_border, num_fmt="#,##0")
        _sc(ws, R, 6, value=summary["total_net_weight"], font=_fnt(11, bold=True),
            align=_aln("right"), border=total_border, num_fmt="#,##0.00")
        _sc(ws, R, 7, value=summary["total_gross_weight"], font=_fnt(11, bold=True),
            align=_aln("right"), border=total_border, num_fmt="#,##0.00")
        R += 1

        # Footer
        R += 1
        _sc(ws, R, 1, value=f"PACKED IN {summary['total_pallets']} PALLETS ONLY.",
            font=_fnt(10))
        R += 1
        _sc(ws, R, 1, value=f"TOTAL MEASUREMENT: {summary['total_measurement_m3']:.2f} m³",
            font=_fnt(10))
        R += 1
        _sc(ws, R, 1, value="PACKING: BULK IN CARTON, THEN ON EURO-PALLET.",
            font=_fnt(10))
        R += 1
        R += 1
        _sc(ws, R, 1, value=f"N.W.: {summary['total_net_weight']:,.2f} KGS   "
                             f"G.W.: {summary['total_gross_weight']:,.2f} KGS",
            font=_fnt(10, bold=True))

        wb.save(output_path)

        # Write back to model
        model.derived.total_cartons = summary["total_cartons"]
        model.derived.pallet_count = summary["total_pallets"]
        model.derived.total_net_weight = summary["total_net_weight"]
        model.derived.total_gross_weight = summary["total_gross_weight"]
        model.derived.total_measurement_m3 = summary["total_measurement_m3"]

        return {
            "success": True,
            "pl_path": output_path,
            "items": len(packing_lines),
            "total_cartons": summary["total_cartons"],
            "total_pallets": summary["total_pallets"],
            "total_net_weight": summary["total_net_weight"],
            "total_gross_weight": summary["total_gross_weight"],
        }
