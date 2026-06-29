"""
writers/base_writer.py — Writer 公共基类 + 共享辅助函数

所有 Writer（Quote/PI/CI/PL）继承此基类，保持统一接口。
共享的 openpyxl cell/merge/border 辅助函数也集中在此。
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

FONT_NAME = "Calibri"


def sc(ws, r, c, value=None, font=None, align=None, border=None, num_fmt=None):
    """Set cell: write value and apply styles to a single cell."""
    cell = ws.cell(row=r, column=c)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def mc(ws, r1, c1, r2, c2, value=None, font=None, align=None):
    """Merge cells and set value/style on the top-left cell."""
    if not (r1 == r2 and c1 == c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    return cell


def brd(top=None, bottom=None, left=None, right=None):
    """Create a Border with named side styles (e.g. 'thin', 'medium')."""
    def _side(x):
        return Side(style=x) if x else Side(style=None)
    return Border(top=_side(top), bottom=_side(bottom), left=_side(left), right=_side(right))


def rh(ws, r, pts):
    """Set row height in points."""
    ws.row_dimensions[r].height = pts


class BaseWriter:
    """Writer 统一基类"""

    def __init__(self, model, config: dict | None = None):
        """
        参数:
            model: OrderModel 实例（必须已 resolve）
            config: 完整 config dict（PLWriter 需要）
        """
        self.model = model
        self.config = config or {}

        if model.resolved:
            self.seller = model.resolved.seller
            self.buyer = model.resolved.buyer
            self.terms = model.resolved.terms
        else:
            self.seller = {}
            self.buyer = {}
            self.terms = {}

    def write(self, output_path: str, **kwargs) -> dict:
        """
        生成文档。子类必须实现。

        返回:
            dict — 生成结果摘要
        """
        raise NotImplementedError(f"{self.__class__.__name__}.write() 未实现")

    # ── 样式辅助方法 ──

    @staticmethod
    def fnt(size=10, bold=False, color="222222"):
        return Font(name=FONT_NAME, size=size, bold=bold, color=color)

    @staticmethod
    def aln(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    @staticmethod
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    @staticmethod
    def thin(color="E0E0E0"):
        return Side(style="thin", color=color)

    @staticmethod
    def med(color="1F3864"):
        return Side(style="medium", color=color)
