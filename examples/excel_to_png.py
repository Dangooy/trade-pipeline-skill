"""Convert sample Excel files to high-resolution PNG screenshots."""
from pathlib import Path

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    print("Pillow not installed. Run: pip install Pillow")
    exit(1)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FONT_PATH = "C:/Windows/Fonts/calibri.ttf"
FONT_PATH_BOLD = "C:/Windows/Fonts/calibrib.ttf"
FONT_PATH_CN = "C:/Windows/Fonts/simsun.ttc"

SCALE = 2.0


def get_font(size=16, bold=False):
    try:
        path = FONT_PATH_BOLD if bold else FONT_PATH
        return ImageFont.truetype(path, int(size * SCALE))
    except Exception:
        return ImageFont.load_default()


def get_cn_font(size=16, bold=True):
    try:
        return ImageFont.truetype(FONT_PATH_CN, int(size * SCALE))
    except Exception:
        return get_font(size, bold)


def hex_to_rgb(hex_color):
    if not hex_color or hex_color == "00000000" or len(hex_color) < 6:
        return None
    h = hex_color[-6:]
    try:
        return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))
    except Exception:
        return None


def render_excel_to_png(xlsx_path, out_path, max_rows=30, max_cols=None):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    if max_cols is None:
        max_cols = min(ws.max_column, 18)
    nrows = min(ws.max_row, max_rows)

    col_widths_px = []
    for c in range(1, max_cols + 1):
        letter = get_column_letter(c)
        w = ws.column_dimensions[letter].width or 8.43
        col_widths_px.append(max(int(w * 10 * SCALE), int(40 * SCALE)))

    row_height = int(28 * SCALE)
    pad_x, pad_y = int(20 * SCALE), int(15 * SCALE)
    total_w = sum(col_widths_px) + pad_x * 2
    total_h = nrows * row_height + pad_y * 2

    img = Image.new("RGB", (total_w, total_h), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    y = pad_y
    for r in range(1, nrows + 1):
        x = pad_x
        for c in range(1, max_cols + 1):
            cell = ws.cell(r, c)
            cw = col_widths_px[c - 1]

            bg = None
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                bg = hex_to_rgb(str(cell.fill.fgColor.rgb))
            if bg:
                draw.rectangle([x, y, x + cw, y + row_height], fill=bg)

            val = cell.value
            if val is None:
                val = ""
            elif isinstance(val, float):
                val = f"{val:,.2f}"
            elif isinstance(val, int):
                val = f"{val:,}"
            else:
                val = str(val).replace("\n", " ")

            if len(val) > 40:
                val = val[:38] + ".."

            is_bold = cell.font and cell.font.bold
            font_color = (0, 0, 0)
            if cell.font and cell.font.color and cell.font.color.rgb:
                fc = hex_to_rgb(str(cell.font.color.rgb))
                if fc:
                    font_color = fc

            font_size = 14
            if cell.font and cell.font.size:
                font_size = min(int(cell.font.size), 18)

            has_cn = any('一' <= ch <= '鿿' for ch in val)
            if has_cn:
                font = get_cn_font(font_size, is_bold)
            else:
                font = get_font(font_size, is_bold)

            text_y = y + int(4 * SCALE)
            draw.text((x + int(5 * SCALE), text_y), val, fill=font_color, font=font)

            x += cw
        y += row_height

    draw.rectangle(
        [pad_x - 1, pad_y - 1, total_w - pad_x, total_h - pad_y],
        outline=(200, 200, 200), width=int(2 * SCALE),
    )

    img.save(out_path, "PNG", dpi=(144, 144))
    print(f"  Saved: {out_path} ({total_w}x{total_h})")
    wb.close()


sample_dir = Path(__file__).resolve().parent / "sample_output"
docs_dir = Path(__file__).resolve().parent.parent / "docs"

for name, max_c, max_r in [
    ("2601_quotation.xlsx", 8, 32),
    ("2601_pi.xlsx", 15, 35),
    ("2601_ci.xlsx", 17, 35),
    ("2601_pl.xlsx", 7, 30),
]:
    src = sample_dir / name
    if src.exists():
        out = docs_dir / name.replace(".xlsx", ".png")
        print(f"Rendering {name}...")
        render_excel_to_png(str(src), str(out), max_rows=max_r, max_cols=max_c)

print("\nDone!")
