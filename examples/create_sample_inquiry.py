"""Generate a sample inquiry Excel file for demonstration."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Inquiry"

header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
thin = Side(style="thin", color="D0D0D0")
border = Border(bottom=thin)

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 6
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 10

ws.merge_cells("A1:G1")
ws.cell(1, 1, "INQUIRY — Global Fasteners LLC").font = Font(size=16, bold=True)

headers = ["No.", "Barcode", "Commodity Description", "UOM", "Total Quantity (PCS)", "KG/MPCS", "Quantity in Box"]
for i, h in enumerate(headers, 1):
    c = ws.cell(3, i, h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", wrap_text=True)

items = [
    (1, "GF-HB-001", "HEX HEAD BOLT DIN 933 M8×25 ZP", "pcs", 50000, 5.80, 200),
    (2, "GF-HB-002", "HEX HEAD BOLT DIN 933 M10×30 ZP", "pcs", 30000, 11.20, 100),
    (3, "GF-HB-003", "HEX HEAD BOLT DIN 933 M12×40 ZP", "pcs", 20000, 20.50, 50),
    (4, "GF-HN-001", "HEX NUT DIN 934 M8 ZP", "pcs", 100000, 2.80, 500),
    (5, "GF-HN-002", "HEX NUT DIN 934 M10 ZP", "pcs", 80000, 5.60, 300),
    (6, "GF-HN-003", "HEX NUT DIN 934 M12 ZP", "pcs", 60000, 9.20, 200),
    (7, "GF-FW-001", "FLAT WASHER DIN 125 M8 ZP", "pcs", 200000, 1.50, 1000),
    (8, "GF-FW-002", "FLAT WASHER DIN 125 M10 ZP", "pcs", 150000, 2.90, 800),
    (9, "GF-FW-003", "FLAT WASHER DIN 125 M12 ZP", "pcs", 100000, 4.80, 500),
    (10, "GF-SW-001", "SPRING LOCK WASHER DIN 127 M8 ZP", "pcs", 100000, 1.20, 1000),
]

for row_data in items:
    r = row_data[0] + 3
    for i, v in enumerate(row_data, 1):
        c = ws.cell(r, i, v)
        c.font = Font(name="Calibri", size=10)
        c.border = border
        if i >= 5:
            c.number_format = "#,##0" if i == 5 else "#,##0.00"

out = "sample_inquiry.xlsx"
wb.save(out)
print(f"Created: {out}")
