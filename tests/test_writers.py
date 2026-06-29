"""Smoke tests for Writer modules (Quote/PI/CI/PL)."""
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from tests.conftest import make_resolved_model, SAMPLE_CONFIG
from trade_pipeline.writers.quote_writer import QuoteWriter
from trade_pipeline.writers.pi_writer import PIWriter
from trade_pipeline.writers.ci_writer import CIWriter, amount_to_words
from trade_pipeline.writers.pl_writer_lite import PLWriterLite


# ── QuoteWriter ──────────────────────────────────────────────────


def test_quote_generates_valid_xlsx():
    model = make_resolved_model()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        result = QuoteWriter(model, SAMPLE_CONFIG).write(path)

        assert Path(path).exists()
        assert result["items"] == 2
        assert result["uuid_col"] == 9
        assert result["uuid_col_letter"] == "I"
        assert result["data_start_row"] > 1

        wb = load_workbook(path)
        ws = wb.active
        assert ws.title == "Quotation"
        assert ws.column_dimensions["I"].hidden is True

        uuid_vals = [ws.cell(r, 9).value for r in range(result["data_start_row"],
                     result["data_start_row"] + result["items"])]
        assert "abc123def456" in uuid_vals
        assert "xyz789ghi012" in uuid_vals
    finally:
        Path(path).unlink(missing_ok=True)


def test_quote_price_column_highlighted():
    model = make_resolved_model()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        result = QuoteWriter(model, SAMPLE_CONFIG).write(path)
        wb = load_workbook(path)
        ws = wb.active
        price_cell = ws.cell(result["data_start_row"], 6)
        assert price_cell.fill.fgColor.rgb is not None
    finally:
        Path(path).unlink(missing_ok=True)


# ── PIWriter ─────────────────────────────────────────────────────


def test_pi_generates_with_correct_structure():
    model = make_resolved_model(with_prices=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        result = PIWriter(model, SAMPLE_CONFIG).write(path)

        assert Path(path).exists()
        assert result["pi_number"] == "PI-TEST01"
        assert result["items"] == 2

        wb = load_workbook(path)
        ws = wb.active
        assert ws.title == "PI"
    finally:
        Path(path).unlink(missing_ok=True)


# ── CIWriter ─────────────────────────────────────────────────────


def test_ci_generates_with_total_amount():
    model = make_resolved_model(with_prices=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        result = CIWriter(model, SAMPLE_CONFIG).write(path)

        assert Path(path).exists()
        assert result["ci_number"] == "CI-TEST01"
        expected_total = 28.50 * 50000 + 12.00 * 100000
        assert result["total_amount"] == expected_total

        wb = load_workbook(path)
        ws = wb.active
        assert ws.title == "CI-TEST01"
    finally:
        Path(path).unlink(missing_ok=True)


# ── amount_to_words ──────────────────────────────────────────────


def test_amount_to_words_zero():
    assert amount_to_words(0, "USD") == "USD ZERO ONLY"


def test_amount_to_words_with_cents():
    result = amount_to_words(1234.56, "USD")
    assert "ONE THOUSAND" in result
    assert "TWO HUNDRED" in result
    assert "THIRTY" in result
    assert "FOUR" in result
    assert "FIFTY" in result
    assert "SIX CENTS" in result
    assert result.endswith("ONLY")


def test_amount_to_words_million():
    result = amount_to_words(1000000, "EUR")
    assert "EUR" in result
    assert "ONE MILLION" in result
    assert result.endswith("ONLY")


def test_amount_to_words_whole_number():
    result = amount_to_words(500, "USD")
    assert "FIVE HUNDRED" in result
    assert "CENTS" not in result
    assert result.endswith("ONLY")


# ── PLWriterLite ─────────────────────────────────────────────────


def test_pl_lite_computes_packing():
    model = make_resolved_model(with_weights=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        result = PLWriterLite(model, SAMPLE_CONFIG).write(path)

        assert Path(path).exists()
        assert result["success"] is True
        assert result["total_cartons"] > 0
        assert result["total_pallets"] > 0
        assert result["total_net_weight"] > 0

        wb = load_workbook(path)
        ws = wb.active
        assert ws.title == "PACKING LIST"
    finally:
        Path(path).unlink(missing_ok=True)


def test_pl_raises_when_weights_missing():
    """Safety net: PL must refuse to generate when no weight info available."""
    import pytest
    from trade_pipeline.writers.pl_writer_lite import PackingInfoMissingError

    model = make_resolved_model(with_weights=False)  # no weight_kg, no kg_mpcs
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        with pytest.raises(PackingInfoMissingError) as exc_info:
            PLWriterLite(model, SAMPLE_CONFIG).write(path)
        assert len(exc_info.value.missing_items) == 2
        assert "缺少重量信息" in str(exc_info.value)
    finally:
        Path(path).unlink(missing_ok=True)


def test_pl_allow_missing_weight_flag_bypasses_check():
    """Escape hatch: explicit allow_missing_weight=True still generates (for demos)."""
    model = make_resolved_model(with_weights=False)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        result = PLWriterLite(model, SAMPLE_CONFIG).write(path, allow_missing_weight=True)
        assert result["success"] is True
        assert Path(path).exists()
    finally:
        Path(path).unlink(missing_ok=True)
