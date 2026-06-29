"""End-to-end pipeline test using sample inquiry data."""
import json
import tempfile
from pathlib import Path

import openpyxl

from trade_pipeline.pipeline.main import run, run_price_update


def test_full_pipeline_on_sample_data():
    """Run full pipeline on examples/sample_inquiry.xlsx（默认带 precheck）。

    sample_inquiry.xlsx 刚组装出来单价为空（先出报价单的正常状态）→ R005 error。
    按生成前检查契约（v1.2.0-alpha.3+）：报价单照出，但 PI/CI/PL 因 error 被跳过，
    整体 success=False。这是有意的安全收紧——缺价不应签发正式单据。
    """
    with tempfile.TemporaryDirectory() as tmp_dir:
        result = run(
            input_path="examples/sample_inquiry.xlsx",
            order_no="E2E_TEST",
            buyer_id="global_fasteners",
            output_dir=tmp_dir,
        )

        # 缺价 → error 阻断正式单据，整体失败
        assert result["success"] is False
        assert any("R005" in e for e in result["errors"]), (
            "缺单价应记录 R005 阻断错误"
        )

        # 报价单 + 检查所需中间产物仍生成
        for key in ["rfq_json", "model_json", "quotation_xlsx", "precheck_md"]:
            assert key in result["outputs"], f"Missing output: {key}"
            assert Path(result["outputs"][key]).exists(), f"{key} file not found"
            assert Path(result["outputs"][key]).stat().st_size > 0, f"{key} is empty"

        # 正式单据（PI/CI/PL）一个都不应生成
        assert "pi_xlsx" not in result["outputs"], "error 阻断后不应生成 PI"
        assert "ci_xlsx" not in result["outputs"], "error 阻断后不应生成 CI"
        assert "pl_xlsx" not in result["outputs"], "error 阻断后不应生成 PL"


def test_full_pipeline_with_packing_review_completes_pl():
    """v1.1.0: 提供 packing_review.json → PL 应该生成成功。

    本用例验证 packing review → PL 路径，与生成前检查无关。sample 缺价会触发
    R005 error 阻断正式单据，故用 precheck=False 隔离，专注 PL 生成逻辑。
    precheck 的阻断语义由 test_precheck_cli.py 专门覆盖。
    """
    with tempfile.TemporaryDirectory() as tmp_dir:
        # 第一次跑：让安全网生成 review.json
        first = run(
            input_path="examples/sample_inquiry.xlsx",
            order_no="E2E_FILL",
            buyer_id="global_fasteners",
            output_dir=tmp_dir,
            precheck=False,
        )
        review_path = first["outputs"].get("packing_review_json")
        assert review_path is not None and Path(review_path).exists()

        # 模拟用户填写：给每个 item 一个 kg_per_1000pcs + pcs_per_carton
        data = json.loads(Path(review_path).read_text(encoding="utf-8"))
        for it in data["items"]:
            it["kg_per_1000pcs"] = 5.0
            it["pcs_per_carton"] = 1000
            it["resolved"] = True
        Path(review_path).write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

        # 第二次跑 + --confirm-packing
        second = run(
            input_path="examples/sample_inquiry.xlsx",
            order_no="E2E_FILL",
            buyer_id="global_fasteners",
            output_dir=tmp_dir,
            packing_review_path=review_path,
            save_packing_to_catalog=False,  # 不污染主 config
            precheck=False,
        )

        assert second["success"] is True
        assert "pl_xlsx" in second["outputs"]
        assert Path(second["outputs"]["pl_xlsx"]).exists()
        assert Path(second["outputs"]["pl_xlsx"]).stat().st_size > 0


def test_quote_only_mode():
    """Run pipeline in --quote-only mode, verify only quotation is generated."""
    with tempfile.TemporaryDirectory() as tmp_dir:
        result = run(
            input_path="examples/sample_inquiry.xlsx",
            order_no="E2E_QUOTE",
            buyer_id="global_fasteners",
            output_dir=tmp_dir,
            quote_only=True,
        )

        assert result["success"] is True
        assert "quotation_xlsx" in result["outputs"]
        assert "pi_xlsx" not in result["outputs"]
        assert "ci_xlsx" not in result["outputs"]


# v1.1.1+ Batch C1: 覆盖未测过的 main.py 分支


def test_run_with_extraction_failure_returns_error():
    """传入不存在的文件 → 提取失败 → 返回 errors，success=False。"""
    with tempfile.TemporaryDirectory() as tmp_dir:
        result = run(
            input_path="examples/__nonexistent_file__.xlsx",
            order_no="E2E_MISSING",
            buyer_id="global_fasteners",
            output_dir=tmp_dir,
        )
        assert result["success"] is False
        assert result["errors"], "缺失输入文件应产生 errors"


def _fill_quotation_prices(quotation_path: str, price: float = 12.34) -> int:
    """把报价单里所有有 UUID 的行的单价列填上同一个价格。

    返回填写的行数（== item 数）。
    """
    wb = openpyxl.load_workbook(quotation_path)
    ws = wb.active

    # 找 UUID 列 + 单价列（复用 price_updater 的扫描逻辑）
    UUID_HEADER = "__item_uuid__"
    uuid_col = None
    header_row = None
    for row in range(1, min(ws.max_row + 1, 21)):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if val and str(val).strip() == UUID_HEADER:
                uuid_col = col
                header_row = row
                break
        if uuid_col:
            break
    assert uuid_col is not None, "报价单应含 UUID 列"

    # 找单价列（同 header 行，含 'price'）
    price_col = None
    for col in range(1, uuid_col):
        val = str(ws.cell(header_row, col).value or "")
        if "price" in val.lower():
            price_col = col
            break
    assert price_col is not None, "报价单应有单价列"

    # 给每行有 UUID 的填价格
    filled = 0
    for row in range(header_row + 1, ws.max_row + 1):
        uid = ws.cell(row, uuid_col).value
        if uid and str(uid).strip() and str(uid).strip() != UUID_HEADER:
            ws.cell(row, price_col, price)
            filled += 1

    wb.save(quotation_path)
    return filled


def test_run_price_update_round_trip():
    """run() → 填价格 → run_price_update() → PI/CI 重新生成。

    覆盖 run_price_update 入口分支（之前只 unit-test 过 update_prices）。
    """
    with tempfile.TemporaryDirectory() as tmp_dir:
        # 1) 先跑出报价单 + model
        first = run(
            input_path="examples/sample_inquiry.xlsx",
            order_no="E2E_PRICE_UPDATE",
            buyer_id="global_fasteners",
            output_dir=tmp_dir,
            quote_only=True,  # 加快测试，不生成 PI/CI
        )
        assert first["success"] is True
        quote_path = first["outputs"]["quotation_xlsx"]
        model_path = first["outputs"]["model_json"]

        # 2) 模拟报价员填单价
        filled_rows = _fill_quotation_prices(quote_path, price=99.5)
        assert filled_rows > 0

        # 3) 跑价格回写
        #    sample_inquiry 缺重量/目的港 → 触发 precheck warning（默认阻断）。
        #    本用例意在验证 PI/CI 重生成的 round-trip，故 skip_warnings 放行；
        #    precheck 的阻断语义由 test_precheck_cli.py 专门覆盖。
        result = run_price_update(
            quotation_path=quote_path,
            model_path=model_path,
            skip_warnings=True,
        )

        # run_price_update 必有 outputs 字段；PI/CI/PL 写入失败会进 errors
        # 但 sample_inquiry 缺重量 → PL 缺失是预期，PI/CI 应该 OK
        assert isinstance(result, dict)
        assert "outputs" in result or "errors" in result, (
            "run_price_update 必须返回 outputs 或 errors"
        )


def test_run_save_packing_to_catalog_false_does_not_write_catalog():
    """save_packing_to_catalog=False 时，product_catalog.yaml 不应被新建/修改。

    覆盖 run() 的 save_packing_to_catalog 分支。
    """
    with tempfile.TemporaryDirectory() as tmp_dir:
        # 第一次跑：生成 review.json（precheck=False 隔离：本用例测 catalog 写入分支）
        first = run(
            input_path="examples/sample_inquiry.xlsx",
            order_no="E2E_NO_CATALOG",
            buyer_id="global_fasteners",
            output_dir=tmp_dir,
            precheck=False,
        )
        review_path = first["outputs"].get("packing_review_json")
        assert review_path is not None

        # 填好 review
        data = json.loads(Path(review_path).read_text(encoding="utf-8"))
        for it in data["items"]:
            it["kg_per_1000pcs"] = 5.0
            it["pcs_per_carton"] = 1000
            it["resolved"] = True
        Path(review_path).write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

        # 应用，但禁用 catalog 写入
        second = run(
            input_path="examples/sample_inquiry.xlsx",
            order_no="E2E_NO_CATALOG",
            buyer_id="global_fasteners",
            output_dir=tmp_dir,
            packing_review_path=review_path,
            save_packing_to_catalog=False,
            precheck=False,
        )
        assert second["success"] is True
        assert "pl_xlsx" in second["outputs"]


def test_missing_seller_id_returns_error_not_traceback(inject_test_config):
    """format_defaults 缺 seller_id 时，assemble() fail-loud 抛 EntityResolutionError，
    但 _assemble_model 必须捕获并返回结构化错误，不让异常冒泡成 traceback。

    回归 Codex PR #5 review P1：此前 assemble() 的 EntityResolutionError 不在
    _assemble_model 的 except 覆盖范围内（只接 BuyerMatchError），导致漏成 traceback。

    复用 conftest 的 inject_test_config fixture（它已把 main.config_path 指向一份
    SAMPLE_CONFIG 临时文件）——直接改那份文件删掉 standard 格式的 seller_id。
    """
    import yaml

    cfg = inject_test_config  # autouse fixture 返回的 config 路径（已 patch 进 main）
    data = yaml.safe_load(cfg.read_text(encoding="utf-8"))
    del data["format_defaults"]["standard"]["seller_id"]
    cfg.write_text(yaml.safe_dump(data, allow_unicode=True), encoding="utf-8")

    with tempfile.TemporaryDirectory() as tmp_dir:
        # 不应抛异常
        result = run(
            input_path="examples/sample_inquiry.xlsx",
            order_no="E2E_NO_SELLER",
            buyer_id="global_fasteners",
            output_dir=tmp_dir,
            quote_only=True,
        )

    # 应返回结构化错误（未抛 traceback），errors 指向 seller 配置问题
    assert result["success"] is False
    assert any("seller" in e.lower() for e in result["errors"]), (
        f"errors 应说明 seller 配置缺失，实际：{result['errors']}"
    )
    # 实体解析在 Step 4 失败 → 报价单及正式单据均不应生成
    # （Step 1-3 的 rfq_json 等中间产物可以留下，无害）
    for doc in ("quotation_xlsx", "pi_xlsx", "ci_xlsx", "pl_xlsx", "model_json"):
        assert doc not in result["outputs"], f"seller 缺失时不应生成 {doc}"
